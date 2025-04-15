import sqlite3
import requests
from pathlib import Path
from datetime import datetime
import tempfile
import os
from openpyxl import load_workbook

# Configuration
EXCEL_URL = "https://cta-service-cms2.hubspot.com/web-interactives/public/v1/track/click?encryptedPayload=AVxigLIgYK73Kr%2BiwqkEFezeiIeiGJLo3Hv1qByiMRfvhJRblQQPu22drY9UF7VKzRsbv2Nz2PW9hsme%2BhEGnvPKOSkBFBPm75zSQRam%2BIdqp%2B0QXKMvOsMvAzRolJK3G2aH87tNQAFt2P%2BRTFLeeheF4gtV3cF7YDyEEuzgBp4Zq7SUnPXFXYte&portalId=19621209&webInteractiveId=317337472610&webInteractiveContentId=162977110893&containerType=EMBEDDED&pageUrl=https%3A%2F%2Fwww.pensford.com%2Fresources%2Fforward-curve&pageTitle=Forward+Curve&referrer=&userAgent=Mozilla%2F5.0+%28Macintosh%3B+Intel+Mac+OS+X+10_15_7%29+AppleWebKit%2F605.1.15+%28KHTML%2C+like+Gecko%29+Version%2F15.6.1+Safari%2F605.1.15&hutk=835f7f3168c74c820841f5d26cbe5af2&hssc=162743459.3.1744371838504&hstc=162743459.835f7f3168c74c820841f5d26cbe5af2.1744316769000.1744370575003.1744371838504.3&pageId=59499436115&analyticsPageId=59499436115&hsfp=2568952849&canonicalUrl=https%3A%2F%2Fwww.pensford.com%2Fresources%2Fforward-curve&contentType=blog-post"
DB_DIR = Path("../data")
DB_FILE = DB_DIR / "rates.db"

def download_excel_file(url):
    """Download the Excel file with the current rates from Penford."""
    try:
        print("Downloading Excel file from Penford")
        # Attempt to download the file from the URL
        response = requests.get(url, timeout=10)
        response.raise_for_status() 
        # Save the file in a temporary location
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_file.write(response.content)
            temp_file_path = temp_file.name
        print(f"Excel file downloaded to temporary path: {temp_file_path}")
        return temp_file_path
    except requests.RequestException as e:
        raise Exception(f"Failed to download Excel file: {e}")

def process_forward_curve(temp_file_path):
    """Read and process the forward curve Excel data, and stop when empty row found"""
    try:
        print("Processing Excel file")
        # Load the file and read the data
        wb = load_workbook(temp_file_path, data_only=True)
        ws = wb.active
         
        data = [] # Variable to store the data extracted
        
        for row_idx in range(6, ws.max_row + 1): # Data in excel file starts from Row 6
            reset_date = ws.cell(row=row_idx, column=7).value # Reset Date in Column 7
            sofr_rate = ws.cell(row=row_idx, column=8).value # SOFR Rates in Column 8
            
            # Stop when empty row found
            if reset_date is None and sofr_rate is None:
                print(f"Stopping at row {row_idx}: invalid row detected")
                break
            
            # Check the RESET DATE and configure it
            try:
                date_str = reset_date.strftime("%Y-%m-%d")
            except ValueError:
                    print(f"Row {row_idx}: G{row_idx} '{reset_date}' is not a valid date")
                    date_str = None
            
            # Check if MARKET EXPECTATIONS is numeric
            try:
                sofr_rate = float(sofr_rate)
            except (TypeError, ValueError):
                print(f"Row {row_idx}: H{row_idx} '{sofr_rate}' is not numeric")
                sofr_rate = None
            
            # Store the data, if both columns are valid
            if date_str is not None and sofr_rate is not None:
                data.append({"Date": date_str, "SOFR_Rate": sofr_rate})
        
        # If no data found
        if not data:
            raise ValueError("No valid data found.")
        
        # Return the processed data
        print(f"Processed {len(data)} valid records.")
        return data
    
    except Exception as e:
        raise Exception(f"Error processing forward curve data: {e}")

def load_to_database(data, db_file):
    """Load the processed data into a SQLite database."""

    DB_DIR.mkdir(exist_ok=True)
    try:
        # Create SQL connection
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        
        # Create table to store the data
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS sofr_rates (
                date TEXT PRIMARY KEY,
                rate REAL
            )
        """)

        # Clear the existing data to replace with latest data
        cursor.execute("DELETE FROM sofr_rates")

        # Insert the data in the database table
        for row in data:
            cursor.execute("""
                INSERT INTO sofr_rates (date, rate) VALUES (?, ?)
            """, (row["Date"], row["SOFR_Rate"]))
        
        conn.commit()
        print(f"Loaded {len(data)} records into {db_file}")

    # If Error encountered in database
    except sqlite3.Error as e:
        raise Exception(f"Database error: {e}")
    
    # Close the connection
    finally:
        conn.close()

def main():
    """Run the ETL process."""

    temp_file_path = None 

    try:
        temp_file_path = download_excel_file(EXCEL_URL) # Path for temporary file downloaded from Pensford
        data = process_forward_curve(temp_file_path) # Data processed from the downloaded file
        load_to_database(data, DB_FILE) # Loading the data in the database

        print("ETL process completed successfully.")

    except Exception as e:
        print(f"ETL process failed: {e}")

    finally:
        # Remove the temporary file that is downloaded, as data is saved
        if temp_file_path and os.path.exists(temp_file_path):
            os.remove(temp_file_path)

if __name__ == "__main__":
    main()